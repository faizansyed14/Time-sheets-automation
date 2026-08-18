"""
Employee Excel Importer.

Parses a .xlsx file with TWO sheets (DXB and AUH), normalises the different
header schemas, and upserts every row into all_employee_data.

DXB headers: "Emp ID", "DCO", "Employees Name", "Project",
             "Account Managers Name", "Contact No.", "Email"
AUH headers: "Employee ID", "Full Name", "Project", "Salesman",
             "Mobile Number", "Email ID"
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee

_BATCH_FLUSH = 200

# Columns an import may change on an EXISTING employee. `employee_id` is part
# of the identity key (always equal on a match) and `active` is a deliberate
# admin decision, so a bulk upload never touches either.
_DIFF_FIELDS = (
    "name", "account_manager", "employee_email_id",
    "project", "contact_no", "location", "all_emails",
)
# ACO and DCO arrive in ONE spreadsheet column, so they're diffed as a pair —
# see _changes_for. A row that fills that cell is authoritative for both.
_REF_FIELDS = ("aco_number", "dco_number")


def _norm(s: Any) -> str:
    """Strip whitespace; return empty string for None/NaN."""
    if s is None:
        return ""
    text = str(s).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip()).lower()


def _identity_key(employee_id: str, name: str) -> tuple[str, str]:
    return (employee_id.strip(), _norm_name(name))


_BLANK_REF = {"", "NA", "N/A", "-", "--", "NIL", "NONE"}
# "DCO2409846_YASIN WARAK" / "ACO1808209_MOHAMED ABDULLAH ELIMAM MOHAMED" /
# "DCO-2102158" — prefix, then the number. Anything after it is the employee's
# own name repeated, which we drop.
_REF_RE = re.compile(r"(ACO|DCO)\s*[-_: ]?\s*(\d+)", re.I)


def _split_ref(raw: str) -> tuple[str | None, str | None]:
    """One ACO/DCO cell -> (aco_number, dco_number).

    The real sheet keeps BOTH kinds in a single column named "DCO", telling
    them apart only by the prefix on the value, and appends the employee's
    name: "DCO2409846_YASIN WARAK", "ACO1808209_MOHAMED ABDULLAH ...". So the
    prefix decides which column the number belongs in, and the trailing name is
    dropped — the vault folder already leads with the person's name, and
    "(DCO-2409846_YASIN WARAK)" would just repeat it.

    A bare number with no prefix is treated as a DCO (that's the column's own
    name). "N/A" and friends mean no number at all.
    """
    v = (raw or "").strip()
    if v.upper().replace(" ", "") in _BLANK_REF:
        return None, None
    m = _REF_RE.search(v)
    if m:
        num = m.group(2)
        return (num, None) if m.group(1).upper() == "ACO" else (None, num)
    digits = re.sub(r"\D", "", v.split("_")[0])
    return (None, digits or None)


def _first_email(raw: str) -> str | None:
    if not raw:
        return None
    for addr in re.split(r"[;,]", raw):
        addr = addr.strip()
        if addr:
            return addr
    return None


def _parse_sheet_dxb(ws) -> list[dict]:
    header_row_num = None
    header_idx = {}
    for i, row in enumerate(ws.iter_rows()):
        cells = [_norm(c.value) for c in row]
        if "Emp ID" in cells or "emp id" in [c.lower() for c in cells]:
            header_row_num = i + 1
            for j, h in enumerate(cells):
                header_idx[h.lower().strip()] = j
            break
    if header_row_num is None:
        return []

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
        row_num = header_row_num + 1 + i
        cells = [_norm(c.value) for c in row]
        if all(c == "" for c in cells):
            continue

        def gl(keys: list[str]) -> str:
            for k in keys:
                v = cells[header_idx[k]] if k in header_idx and header_idx[k] < len(cells) else ""
                if v:
                    return v
            return ""

        emp_id = gl(["emp id"])
        emp_name = gl(["employees name"])
        # Nothing identifiable at all = a spacer/layout row; ignore it quietly.
        # A row with only ONE of the two IS a data-entry problem, so it's kept
        # and reported as skipped rather than vanishing without trace.
        if not emp_id and not emp_name:
            continue
        aco, dco = _split_ref(gl(["dco", "aco/dco", "aco / dco", "dco number",
                                  "dco no.", "dco no", "aco", "reference"]))
        all_emails_raw = gl(["email"])
        records.append({
            "employee_id": emp_id,
            "name": emp_name,
            "aco_number": aco,
            "dco_number": dco,
            # This row filled the ACO/DCO cell, so it decides BOTH fields —
            # an ACO value must also clear a stale DCO (see _changes_for).
            "_ref_present": bool(aco or dco),
            "project": gl(["project"]),
            "account_manager": gl(["account managers name"]),
            "contact_no": gl(["contact no."]),
            "all_emails": all_emails_raw,
            "employee_email_id": _first_email(all_emails_raw),
            "location": "DXB",
            "_row": row_num,
            "_sheet": ws.title,
        })
    return records


def _parse_sheet_auh(ws) -> list[dict]:
    header_idx: dict[str, int] = {}
    header_row_num = None
    for i, row in enumerate(ws.iter_rows()):
        cells = [_norm(c.value) for c in row]
        low = [c.lower() for c in cells]
        if "employee id" in low or "full name" in low:
            header_row_num = i + 1
            for j, h in enumerate(low):
                header_idx[h.strip()] = j
            break
    if not header_idx:
        return []

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
        row_num = header_row_num + 1 + i
        cells = [_norm(c.value) for c in row]
        if all(c == "" for c in cells):
            continue

        def gl(keys: list[str]) -> str:
            for k in keys:
                v = cells[header_idx[k]] if k in header_idx and header_idx[k] < len(cells) else ""
                if v:
                    return v
            return ""

        emp_id = gl(["employee id"])
        emp_name = gl(["full name"])
        if not emp_id and not emp_name:
            continue      # spacer/layout row — see the DXB parser above
        all_emails_raw = gl(["email id", "email"])
        aco, dco = _split_ref(gl(["dco", "aco/dco", "aco / dco", "aco", "reference"]))
        records.append({
            "employee_id": emp_id,
            "name": emp_name,
            "aco_number": aco,
            "dco_number": dco,
            "_ref_present": bool(aco or dco),
            "project": gl(["project"]),
            "account_manager": gl(["salesman"]),
            "contact_no": gl(["mobile number", "contact no."]),
            "all_emails": all_emails_raw,
            "employee_email_id": _first_email(all_emails_raw),
            "location": "AUH",
            "_row": row_num,
            "_sheet": ws.title,
        })
    return records


def parse_workbook(data: bytes) -> list[dict]:
    """xlsx bytes -> raw parsed rows (DXB + AUH sheets, normalised headers)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel import.")

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    records: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_lower = sheet_name.strip().lower()
        if "dxb" in name_lower:
            records.extend(_parse_sheet_dxb(ws))
        elif "auh" in name_lower:
            records.extend(_parse_sheet_auh(ws))
        else:
            r = _parse_sheet_dxb(ws)
            if not r:
                r = _parse_sheet_auh(ws)
            records.extend(r)
    return records


def _split_usable(records: list[dict]) -> tuple[list[dict], list[dict]]:
    """(usable rows, skipped rows). Skips a row with no ID/name, and the second
    occurrence of the same (ID + name) WITHIN this file."""
    usable: list[dict] = []
    skipped: list[dict] = []
    seen: set[tuple[str, str]] = set()
    for rec in records:
        emp_id = (rec.get("employee_id") or "").strip()
        emp_name = (rec.get("name") or "").strip()
        base = {"sheet": rec.get("_sheet", "Unknown"), "row": rec.get("_row", 0),
                "id": emp_id, "name": emp_name}
        if not emp_id or not emp_name:
            skipped.append({**base, "reason": "Missing ID or Name"})
            continue
        key = _identity_key(emp_id, emp_name)
        if key in seen:
            skipped.append({**base, "reason": "Duplicate ID + Name in file"})
            continue
        seen.add(key)
        usable.append({**rec, "_key": key, "employee_id": emp_id, "name": emp_name})
    return usable, skipped


def _clean(v) -> str | None:
    if v is None:
        return None
    s = v.strip() if isinstance(v, str) else str(v).strip()
    return s or None


def _changes_for(existing: Employee, rec: dict) -> list[dict]:
    """Fields this row would actually change on an existing employee.

    A BLANK cell is "no information", never an instruction to erase — a
    partial sheet must not wipe an email/phone/project the DB already knows.
    Only a non-empty value that differs from what's stored counts.

    ACO/DCO are the one exception, because they share a single column: a row
    that fills it decides BOTH, so an "ACO…" value also CLEARS a DCO stored
    from an earlier read (and vice versa). Without that, a person whose number
    was re-issued under the other prefix would keep both forever and their
    vault folder would claim two contracts.
    """
    out: list[dict] = []
    if rec.get("_ref_present"):
        for f in _REF_FIELDS:
            new = _clean(rec.get(f))
            old = _clean(getattr(existing, f, None))
            if old != new:
                out.append({"field": f, "old": old, "new": new})
    for f in _DIFF_FIELDS:
        new = _clean(rec.get(f))
        if new is None:
            continue
        old = _clean(getattr(existing, f, None))
        if old != new:
            out.append({"field": f, "old": old, "new": new})
    return out


async def _load_employees(db: AsyncSession) -> list[Employee]:
    return list((await db.execute(select(Employee))).scalars().all())


def _index_by_identity(rows: list[Employee]) -> dict[tuple[str, str], Employee]:
    return {_identity_key(e.employee_id, e.name or ""): e for e in rows}


def _existing_out(e: Employee) -> dict:
    return {
        "id": e.id, "employee_id": e.employee_id, "name": e.name,
        "location": e.location, "account_manager": e.account_manager,
        "employee_email_id": e.employee_email_id, "active": e.active,
    }


async def build_import_plan(db: AsyncSession, data: bytes) -> dict:
    """Dry run: exactly what this file WOULD do, without writing anything.

    Returned so a reviewer can confirm before committing:
      to_add            — not in the matcher yet (with a possible-rename hint)
      to_update         — matched, with the exact field-level old -> new diffs
      unchanged         — matched, nothing to change
      missing_from_file — in the matcher but NOT in this file. Never deleted;
                          surfaced so a leaver can be spotted and marked
                          inactive deliberately rather than by a silent purge.
      skipped           — unusable rows (no ID/name, duplicated in the file)
    """
    usable, skipped = _split_usable(parse_workbook(data))
    rows = await _load_employees(db)
    index = _index_by_identity(rows)

    # Same ID in the same office under a DIFFERENT name is very likely a
    # rename — flagged, never auto-merged: AUH and DXB reuse ID ranges, so
    # matching on ID alone could silently fuse two different people.
    by_id_loc: dict[tuple[str, str], list[Employee]] = {}
    for e in rows:
        by_id_loc.setdefault(
            ((e.employee_id or "").strip(), (e.location or "").strip()), []).append(e)

    to_add: list[dict] = []
    to_update: list[dict] = []
    unchanged: list[dict] = []
    matched: set[str] = set()

    for rec in usable:
        existing = index.get(rec["_key"])
        if existing is not None:
            matched.add(existing.id)
            changes = _changes_for(existing, rec)
            base = {"id": existing.id, "employee_id": existing.employee_id,
                    "name": existing.name, "location": existing.location}
            if changes:
                to_update.append({**base, "changes": changes})
            else:
                unchanged.append(_existing_out(existing))
            continue

        same = by_id_loc.get(
            (rec["employee_id"], (rec.get("location") or "").strip()), [])
        to_add.append({
            "employee_id": rec["employee_id"], "name": rec["name"],
            "location": _clean(rec.get("location")),
            "project": _clean(rec.get("project")),
            "account_manager": _clean(rec.get("account_manager")),
            "employee_email_id": _clean(rec.get("employee_email_id")),
            "contact_no": _clean(rec.get("contact_no")),
            "aco_number": _clean(rec.get("aco_number")),
            "dco_number": _clean(rec.get("dco_number")),
            "sheet": rec.get("_sheet"), "row": rec.get("_row"),
            "possible_rename_of": (
                f"{same[0].name} ({same[0].employee_id})" if same else None),
        })

    missing = sorted(
        (_existing_out(e) for e in rows if e.id not in matched),
        key=lambda r: (r["name"] or "").lower())

    return {
        "to_add": to_add,
        "to_update": to_update,
        "unchanged": unchanged,
        "missing_from_file": missing,
        "skipped": skipped,
    }


async def import_employees_from_bytes(db: AsyncSession, data: bytes) -> dict:
    """Parse xlsx bytes and apply: insert new employees, update changed fields
    on existing ones. Purely additive — an employee absent from the file is
    never deleted or deactivated (see build_import_plan's missing_from_file)."""
    usable, skipped = _split_usable(parse_workbook(data))
    rows = await _load_employees(db)
    index = _index_by_identity(rows)

    inserted = updated = touched = 0
    with db.no_autoflush:
        for rec in usable:
            existing = index.get(rec["_key"])
            if existing is not None:
                changes = _changes_for(existing, rec)
                if not changes:
                    continue
                for c in changes:
                    setattr(existing, c["field"], c["new"])
                updated += 1
            else:
                row = Employee(
                    employee_id=rec["employee_id"],
                    name=rec["name"],
                    aco_number=_clean(rec.get("aco_number")),
                    dco_number=_clean(rec.get("dco_number")),
                    account_manager=_clean(rec.get("account_manager")),
                    employee_email_id=_clean(rec.get("employee_email_id")),
                    project=_clean(rec.get("project")),
                    contact_no=_clean(rec.get("contact_no")),
                    location=_clean(rec.get("location")),
                    all_emails=_clean(rec.get("all_emails")),
                )
                db.add(row)
                index[rec["_key"]] = row
                inserted += 1

            touched += 1
            if touched % _BATCH_FLUSH == 0:
                await db.flush()

    await db.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": len(skipped),
        "skipped_details": skipped,
    }
