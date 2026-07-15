"""Build an XLSX workbook of timesheet leave data for one calendar month."""
from __future__ import annotations

import calendar
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.extraction.validation import BUCKETS

# (record attribute for dates, count header, dates header)
_LEAVE_SPECS: list[tuple[str, str, str]] = [
    ("annual_leave_dates", "Annual Leave Count", "Annual Leave Dates"),
    ("remote_work_dates", "Remote / WFH Count", "Remote / WFH Dates"),
    ("sick_leave_dates", "Sick Leave Count", "Sick Leave Dates"),
    ("maternity_leave_dates", "Maternity Leave Count", "Maternity Leave Dates"),
    ("unpaid_leave_dates", "Unpaid Leave Count", "Unpaid Leave Dates"),
    ("absent_dates", "Absent Count", "Absent Dates"),
    ("public_holiday_dates", "Public Holiday Count", "Public Holiday Dates"),
]

_EMPLOYEE_HEADERS: list[tuple[str, str]] = [
    ("employee_id", "Employee ID"),
    ("employee_name", "Employee Name"),
    ("dco_number", "DCO Number"),
    ("account_manager", "Account Manager"),
    ("location", "Location"),
    ("project", "Project"),
    ("employee_email", "Email"),
    ("contact_no", "Contact"),
]

_META_HEADERS: list[tuple[str, str]] = [
    ("month_label", "Month"),
    ("year", "Year"),
    ("validation_status", "Validation"),
    ("approval_status", "Approval"),
    ("source_file_count", "Source Files"),
]


def _empty_row_dict(employee: Any, month: int, year: int) -> dict[str, Any]:
    return {
        "employee_id": employee.employee_id or "",
        "employee_name": employee.name or "",
        "dco_number": employee.dco_number or "",
        "account_manager": employee.account_manager or "",
        "location": employee.location or "",
        "project": employee.project or "",
        "employee_email": employee.employee_email_id or "",
        "contact_no": employee.contact_no or "",
        "month_label": calendar.month_name[month],
        "year": year,
        "validation_status": "",
        "approval_status": "",
        "source_file_count": 0,
        **{f"{b}_dates": "" for b in BUCKETS},
        **{f"{b}_count": 0 for b in BUCKETS},
    }


def _row_dict(record: Any, employee: Any | None, month: int, year: int) -> dict[str, Any]:
    attr_map = {
        "annual": "annual_leave_dates",
        "remote": "remote_work_dates",
        "sick": "sick_leave_dates",
        "maternity": "maternity_leave_dates",
        "unpaid": "unpaid_leave_dates",
        "absent": "absent_dates",
        "public_holiday": "public_holiday_dates",
    }
    leave_dates = {b: sorted(getattr(record, attr_map[b]) or []) for b in BUCKETS}

    return {
        "employee_id": record.employee_id or "",
        "employee_name": record.employee_name or "",
        "dco_number": record.dco_number or "",
        "account_manager": record.account_manager or "",
        "location": (employee.location if employee else "") or "",
        "project": (employee.project if employee else "") or "",
        "employee_email": (employee.employee_email_id if employee else "") or "",
        "contact_no": (employee.contact_no if employee else "") or "",
        "month_label": calendar.month_name[month],
        "year": year,
        "validation_status": record.validation_status or "",
        "approval_status": record.approval_status or "",
        "source_file_count": record.source_file_count,
        **{f"{b}_dates": ", ".join(leave_dates[b]) for b in BUCKETS},
        **{f"{b}_count": len(leave_dates[b]) for b in BUCKETS},
    }


def build_timesheet_xlsx(rows: list[dict[str, Any]], month: int, year: int) -> bytes:
    """Return XLSX bytes for the given export rows (from `_row_dict`)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"[:31]

    headers: list[str] = (
        [label for _, label in _EMPLOYEE_HEADERS]
        + [label for _, label in _META_HEADERS]
        + [spec[1] for spec in _LEAVE_SPECS]
        + [spec[2] for spec in _LEAVE_SPECS]
    )

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    attr_map = {
        "annual_leave_dates": "annual",
        "remote_work_dates": "remote",
        "sick_leave_dates": "sick",
        "maternity_leave_dates": "maternity",
        "unpaid_leave_dates": "unpaid",
        "absent_dates": "absent",
        "public_holiday_dates": "public_holiday",
    }

    for row_idx, data in enumerate(rows, 2):
        col = 1
        for key, _ in _EMPLOYEE_HEADERS:
            ws.cell(row=row_idx, column=col, value=data.get(key, ""))
            col += 1
        for key, _ in _META_HEADERS:
            ws.cell(row=row_idx, column=col, value=data.get(key, ""))
            col += 1
        for attr, _, _ in _LEAVE_SPECS:
            bucket = attr_map[attr]
            ws.cell(row=row_idx, column=col, value=data.get(f"{bucket}_count", 0))
            col += 1
        for attr, _, _ in _LEAVE_SPECS:
            bucket = attr_map[attr]
            cell = ws.cell(row=row_idx, column=col, value=data.get(f"{bucket}_dates", ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            col += 1

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(rows) + 2, 52)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(48, len(str(val))))
        ws.column_dimensions[letter].width = max(10, min(44, max_len + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def employees_grid_rows(
    employees: list[Any],
    records_by_pk: dict[str, Any],
    month: int,
    year: int,
) -> list[dict[str, Any]]:
    """One export row per matcher employee; empty leave cells when not filed."""
    out: list[dict[str, Any]] = []
    for emp in employees:
        rec = records_by_pk.get(emp.id)
        if rec:
            out.append(_row_dict(rec, emp, month, year))
        else:
            out.append(_empty_row_dict(emp, month, year))
    out.sort(key=lambda r: (r.get("employee_name") or "").lower())
    return out


def records_to_rows(records: list[Any], employees: dict[str, Any], month: int, year: int) -> list[dict[str, Any]]:
    """Legacy helper — prefer `employees_grid_rows`."""
    by_pk = {r.matched_employee_pk: r for r in records if r.matched_employee_pk}
    emps = list(employees.values()) if employees else []
    if not emps:
        emps = list(by_pk.values())  # fallback: records only
        return [_row_dict(r, employees.get(r.matched_employee_pk or ""), month, year) for r in emps]
    return employees_grid_rows(emps, by_pk, month, year)
