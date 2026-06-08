"""
Employee Excel Importer.

Parses a .xlsx file with TWO sheets (DXB and AUH), normalises the different
header schemas, and upserts every row into all_employee_data.

DXB headers: "Emp ID", "DCO", "Employees Name", "Project",
             "Account Managers Name", "Contact No.", "Email"
AUH headers: "Employee ID", "Full Name", "Project", "Salesman",
             "Mobile Number", "Email ID"
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee


def _norm(s: Any) -> str:
    """Strip whitespace; return empty string for None/NaN."""
    if s is None:
        return ""
    text = str(s).strip()
    # openpyxl sometimes gives float for numeric IDs e.g. 1001.0
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _first_email(raw: str) -> str | None:
    """Return first address from a semicolon- or comma-separated email list."""
    if not raw:
        return None
    for addr in re.split(r"[;,]", raw):
        addr = addr.strip()
        if addr:
            return addr
    return None


def _parse_sheet_dxb(ws) -> list[dict]:
    """Parse the DXB sheet into normalised dicts."""
    # Find the header row (first row that has "Emp ID")
    header_row_num = None
    header_idx = {}
    for i, row in enumerate(ws.iter_rows()):
        cells = [_norm(c.value) for c in row]
        if "Emp ID" in cells or "emp id" in [c.lower() for c in cells]:
            header_row = cells
            header_row_num = i + 1
            for j, h in enumerate(cells):
                header_idx[h.lower().strip()] = j
            break
    if header_row is None:
        return []

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
        row_num = header_row_num + 1 + i
        cells = [_norm(c.value) for c in row]
        if all(c == "" for c in cells):
            continue  # blank row

        def g(key: str) -> str:
            return cells[header_idx[key]] if key in header_idx and header_idx[key] < len(cells) else ""

        # Try lowercased lookups
        def gl(keys: list[str]) -> str:
            for k in keys:
                v = cells[header_idx[k]] if k in header_idx and header_idx[k] < len(cells) else ""
                if v:
                    return v
            return ""

        emp_id = gl(["emp id"])
        if not emp_id:
            continue
        dco_raw = gl(["dco"])
        dco = None if dco_raw.upper() in ("NA", "N/A", "") else dco_raw
        all_emails_raw = gl(["email"])
        records.append({
            "employee_id": emp_id,
            "name": gl(["employees name"]),
            "dco_number": dco,
            "project": gl(["project"]),
            "account_manager": gl(["account managers name"]),
            "contact_no": gl(["contact no."]),
            "all_emails": all_emails_raw,
            "employee_email_id": _first_email(all_emails_raw),
            "location": "DXB",
            "_row": row_num,
            "_sheet": ws.title,
        })
    return records


def _parse_sheet_auh(ws) -> list[dict]:
    """Parse the AUH sheet into normalised dicts."""
    header_idx: dict[str, int] = {}
    header_row_num = None
    for i, row in enumerate(ws.iter_rows()):
        cells = [_norm(c.value) for c in row]
        low = [c.lower() for c in cells]
        if "employee id" in low or "full name" in low:
            header_row_num = i + 1
            for j, h in enumerate(low):
                header_idx[h.strip()] = j
            break
    if not header_idx:
        return []

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
        row_num = header_row_num + 1 + i
        cells = [_norm(c.value) for c in row]
        if all(c == "" for c in cells):
            continue

        def gl(keys: list[str]) -> str:
            for k in keys:
                v = cells[header_idx[k]] if k in header_idx and header_idx[k] < len(cells) else ""
                if v:
                    return v
            return ""

        emp_id = gl(["employee id"])
        if not emp_id:
            continue
        all_emails_raw = gl(["email id", "email"])
        records.append({
            "employee_id": emp_id,
            "name": gl(["full name"]),
            "dco_number": None,
            "project": gl(["project"]),
            "account_manager": gl(["salesman"]),
            "contact_no": gl(["mobile number", "contact no."]),
            "all_emails": all_emails_raw,
            "employee_email_id": _first_email(all_emails_raw),
            "location": "AUH",
            "_row": row_num,
            "_sheet": ws.title,
        })
    return records


async def import_employees_from_bytes(
    db: AsyncSession, data: bytes
) -> dict:
    """
    Parse the xlsx bytes, upsert all rows into all_employee_data.
    Returns {inserted, updated, skipped}.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel import.")

    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    records: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_lower = sheet_name.strip().lower()
        if "dxb" in name_lower:
            records.extend(_parse_sheet_dxb(ws))
        elif "auh" in name_lower:
            records.extend(_parse_sheet_auh(ws))
        else:
            # Try both parsers; take whichever gives results
            r = _parse_sheet_dxb(ws)
            if not r:
                r = _parse_sheet_auh(ws)
            records.extend(r)

    inserted = updated = skipped = 0
    skipped_details = []
    seen_ids: set[str] = set()

    for rec in records:
        emp_id = (rec.get("employee_id") or "").strip()
        emp_name = (rec.get("name") or "").strip()
        
        row_num = rec.get("_row", 0)
        sheet_name = rec.get("_sheet", "Unknown")

        if not emp_id or not emp_name:
            skipped += 1
            skipped_details.append({
                "sheet": sheet_name,
                "row": row_num,
                "id": emp_id,
                "name": emp_name,
                "reason": "Missing ID or Name"
            })
            continue
        if emp_id in seen_ids:
            skipped += 1
            skipped_details.append({
                "sheet": sheet_name,
                "row": row_num,
                "id": emp_id,
                "name": emp_name,
                "reason": "Duplicate ID in file"
            })
            continue
        seen_ids.add(emp_id)

        existing = (
            await db.execute(select(Employee).where(Employee.employee_id == emp_id))
        ).scalar_one_or_none()

        if existing:
            for k, v in rec.items():
                if not k.startswith("_"):
                    setattr(existing, k, v or None)
            updated += 1
        else:
            e = Employee(
                employee_id=emp_id,
                name=emp_name,
                dco_number=rec.get("dco_number"),
                account_manager=rec.get("account_manager") or None,
                employee_email_id=rec.get("employee_email_id") or None,
                project=rec.get("project") or None,
                contact_no=rec.get("contact_no") or None,
                location=rec.get("location"),
                all_emails=rec.get("all_emails") or None,
            )
            db.add(e)
            inserted += 1

    await db.commit()
    return {
        "inserted": inserted, 
        "updated": updated, 
        "skipped": skipped,
        "skipped_details": skipped_details
    }
