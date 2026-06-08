import openpyxl
from io import BytesIO
import os

def _norm(s):
    if s is None:
        return ""
    text = str(s).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text

def check_skips():
    filepath = r"c:\Users\FAIZAN\Downloads\timesheet-portal\backend\app\seed\data\Employee_details.xlsx.xlsx"
    if not os.path.exists(filepath):
        print(f"Error: {filepath} not found.")
        return

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    seen_ids = set()
    skipped_info = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        low_sheet = sheet_name.lower()
        
        # Determine parser type
        is_dxb = "dxb" in low_sheet
        is_auh = "auh" in low_sheet
        
        header_idx = {}
        header_row_num = None
        
        # Find header
        for i, row in enumerate(ws.iter_rows()):
            cells = [_norm(c.value) for c in row]
            low_cells = [c.lower() for c in cells]
            if "emp id" in low_cells or "employee id" in low_cells or "employees name" in low_cells or "full name" in low_cells:
                header_row_num = i + 1
                for j, h in enumerate(low_cells):
                    header_idx[h.strip()] = j
                break
        
        if not header_idx:
            continue

        for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
            cells = [_norm(c.value) for c in row]
            if all(c == "" for c in cells):
                continue
            
            # Extract basic info based on headers found
            emp_id = ""
            if "emp id" in header_idx: emp_id = cells[header_idx["emp id"]]
            elif "employee id" in header_idx: emp_id = cells[header_idx["employee id"]]
            
            emp_name = ""
            if "employees name" in header_idx: emp_name = cells[header_idx["employees name"]]
            elif "full name" in header_idx: emp_name = cells[header_idx["full name"]]

            row_num = header_row_num + 1 + i
            
            reason = ""
            if not emp_id or not emp_name:
                reason = f"Missing ID ({emp_id}) or Name ({emp_name})"
            elif emp_id in seen_ids:
                reason = f"Duplicate ID: {emp_id}"
            else:
                seen_ids.add(emp_id)
                continue # Not skipped
            
            skipped_info.append({
                "sheet": sheet_name,
                "row": row_num,
                "id": emp_id,
                "name": emp_name,
                "reason": reason
            })

    print(f"--- SKIPPED ROWS REPORT ({len(skipped_info)} total) ---")
    for s in skipped_info:
        print(f"Sheet: {s['sheet']} | Row: {s['row']} | ID: {s['id']} | Name: {s['name']} | Reason: {s['reason']}")

if __name__ == "__main__":
    check_skips()
