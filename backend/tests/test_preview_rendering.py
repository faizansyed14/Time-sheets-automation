"""File-preview rendering: real pages, and emails nested inside emails.

Two defects this locks down:

1. Every preview surface reused `to_images`, the VISION renderer, which
   deliberately stitches all pages into one tall canvas (clamped to
   STITCH_MAX_HEIGHT). A multi-page workbook therefore reached the modal as a
   single ~600x8000 image — squeezed to an unreadable smudge, with no page
   navigator because it was one image. Previews now use `to_page_images`.

2. `parse_eml` walked INTO a message/rfc822 part, so a forwarded email arrived
   with size 0 and no filename (unopenable) while its inner parts were hoisted
   into the outer attachment list.
"""
import base64
import io
from email.message import EmailMessage

from app.services.extraction.eml_parser import parse_eml
from app.services.extraction.file_processor import (
    STITCH_MAX_HEIGHT,
    to_images,
    to_page_images,
)


def _multipage_xlsx() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "June"
    for r in range(1, 160):
        for c in range(1, 12):
            ws.cell(r, c, f"R{r}C{c}")
    ws2 = wb.create_sheet("July")
    for r in range(1, 160):
        for c in range(1, 12):
            ws2.cell(r, c, f"X{r}C{c}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sizes(images: list[bytes]) -> list[tuple[int, int]]:
    from PIL import Image

    return [Image.open(io.BytesIO(b)).size for b in images]


def test_multipage_xlsx_previews_as_real_pages():
    """The reported bug: XLSX preview was one unreadable 8000px-tall strip."""
    data = _multipage_xlsx()

    pages = to_page_images("xlsx", data)
    assert len(pages) > 1, "a long workbook must paginate, not collapse to one image"

    sizes = _sizes(pages)
    for w, h in sizes:
        # Nothing like the stitched canvas, and wide enough to read.
        assert h < STITCH_MAX_HEIGHT, f"preview page is a stitched strip: {(w, h)}"
        assert w >= 800, f"preview page too narrow to read: {(w, h)}"


def test_vision_still_gets_one_stitched_image():
    """The LLM path must NOT change — it needs one uninterrupted screenshot so
    no table row is cut across a page break."""
    data = _multipage_xlsx()
    vision = to_images("xlsx", data)
    assert len(vision) == 1
    assert _sizes(vision)[0][1] <= STITCH_MAX_HEIGHT


def test_pdf_preview_paginates_per_page():
    import fitz

    doc = fitz.open()
    for _ in range(3):
        doc.new_page()
    pdf = doc.tobytes()
    doc.close()

    assert len(to_page_images("pdf", pdf)) == 3
    assert len(to_images("pdf", pdf)) == 1   # vision stitches


def _forwarded_email() -> bytes:
    inner = EmailMessage()
    inner["Subject"] = "Re: Approved June 2026 Timesheet"
    inner["From"] = "manager@alpha.ae"
    inner["To"] = "employee@alpha.ae"
    inner.set_content("Approved — see attached.")
    inner.add_attachment(b"%PDF-1.4 inner sheet", maintype="application",
                         subtype="pdf", filename="inner_sheet.pdf")

    outer = EmailMessage()
    outer["Subject"] = "FW: Timesheet June 2026"
    outer["From"] = "employee@alpha.ae"
    outer["To"] = "timesheet@alpha.ae"
    outer.set_content("Forwarding the approval below.")
    # The content manager infers message/rfc822 from the Message object.
    outer.add_attachment(inner)
    return outer.as_bytes()


def test_nested_email_is_kept_whole_and_openable():
    parsed = parse_eml(_forwarded_email())
    nested = [a for a in parsed["attachments"] if a["content_type"] == "message/rfc822"]
    assert len(nested) == 1, parsed["attachments"]

    att = nested[0]
    # Used to be size 0 with filename "attachment" — impossible to open.
    assert att["size"] > 0, "nested email came through empty"
    assert att["filename"].lower().endswith(".eml")
    assert "Approved June 2026" in att["filename"], att["filename"]

    # And it opens as its own email, with its own body and its own files.
    inner = parse_eml(base64.b64decode(att["data_b64"]))
    assert inner["subject"] == "Re: Approved June 2026 Timesheet"
    assert "manager@alpha.ae" in inner["from_"]
    assert [a["filename"] for a in inner["attachments"]] == ["inner_sheet.pdf"]


def test_nested_email_parts_do_not_leak_into_the_outer_email():
    """The forwarded message's own attachment belongs to IT, not to the
    carrier — otherwise the outer email lists files it does not contain."""
    parsed = parse_eml(_forwarded_email())
    names = [a["filename"] for a in parsed["attachments"]]
    assert "inner_sheet.pdf" not in names, names
