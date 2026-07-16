"""Timesheet period export — XLSX download + by-period preview API."""
from io import BytesIO

from openpyxl import load_workbook

from app.models.employee import Employee
from app.models.timesheet_record import TimesheetRecord, ValidationStatus
from tests.conftest import auth_headers


async def test_export_by_period_lists_all_matcher_employees(client, admin_token):
    from app.core.database import SessionLocal

    h = auth_headers(admin_token)
    async with SessionLocal() as db:
        emp = Employee(
            employee_id="E-NO-SUB",
            name="Not Submitted Yet",
            account_manager="Mgr",
            location="DXB",
        )
        db.add(emp)
        await db.commit()
        pk = emp.id

    r = await client.get("/api/v1/timesheets/by-period", params={"month": 3, "year": 2099}, headers=h)
    assert r.status_code == 200, r.text
    row = next((x for x in r.json() if x["employee_id"] == "E-NO-SUB"), None)
    assert row is not None
    assert row["has_record"] is False
    assert row["annual_leave_dates"] == []

    async with SessionLocal() as db:
        obj = await db.get(Employee, pk)
        if obj:
            await db.delete(obj)
            await db.commit()


async def test_export_xlsx_contains_leave_dates(client, admin_token):
    from app.core.database import SessionLocal

    h = auth_headers(admin_token)
    async with SessionLocal() as db:
        emp = Employee(
            employee_id="E-EXP-1",
            name="Export Tester",
            account_manager="Manager A",
            location="AUH",
        )
        db.add(emp)
        await db.flush()
        rec = TimesheetRecord(
            matched_employee_pk=emp.id,
            employee_id="E-EXP-1",
            employee_name="Export Tester",
            account_manager="Manager A",
            month=6,
            year=2026,
            annual_leave_dates=["2026-06-04", "2026-06-05"],
            sick_leave_dates=["2026-06-12"],
            validation_status=ValidationStatus.VERIFIED,
        )
        db.add(rec)
        await db.commit()
        rid = rec.id
        pk = emp.id

    preview = await client.get(
        "/api/v1/timesheets/by-period", params={"month": 6, "year": 2026}, headers=h,
    )
    assert preview.status_code == 200, preview.text
    body = preview.json()
    row = next(r for r in body if r["employee_id"] == "E-EXP-1")
    assert row["id"] == rid
    assert row["has_record"] is True
    assert row["annual_leave_dates"] == ["2026-06-04", "2026-06-05"]
    assert row["sick_leave_dates"] == ["2026-06-12"]

    xlsx = await client.get("/api/v1/timesheets/export", params={"month": 6, "year": 2026}, headers=h)
    assert xlsx.status_code == 200, xlsx.text
    assert "spreadsheetml" in xlsx.headers["content-type"]

    wb = load_workbook(BytesIO(xlsx.content), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[1] == "Export Tester" for r in data_rows)
    tester = next(r for r in data_rows if r[1] == "Export Tester")
    annual_dates_col = headers.index("Annual Leave Dates")
    assert "2026-06-04" in (tester[annual_dates_col] or "")

    async with SessionLocal() as db:
        rec_obj = await db.get(TimesheetRecord, rid)
        if rec_obj:
            await db.delete(rec_obj)
        emp_obj = await db.get(Employee, pk)
        if emp_obj:
            await db.delete(emp_obj)
        await db.commit()
