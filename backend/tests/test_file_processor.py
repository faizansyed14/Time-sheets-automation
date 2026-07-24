"""XLSX -> PDF -> image rendering.

A default/unset column width clips the LAST character of anything longer than
it — measured on a real sheet: '2026-06-01' rendered as '2026-06-0' with the
day digit cut under the grid line, and the vision model then read every leave
date one day off. Not a resolution problem — the column itself was too
narrow. _autofit_xlsx_columns widens columns to fit their content before
LibreOffice ever renders the page.
"""
import io

from openpyxl import Workbook, load_workbook

from app.services.extraction.file_processor import _autofit_xlsx_columns


def _xlsx_bytes(rows: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_columns_are_widened_to_fit_their_longest_cell():
    original = _xlsx_bytes([
        ["Date", "Status"],
        ["2026-06-01", "Present"],
        ["2026-06-02", "Annual Leave"],
    ])
    widened = _autofit_xlsx_columns(original)

    wb = load_workbook(io.BytesIO(widened))
    ws = wb.active
    # "2026-06-01" is 10 chars; a default/unset width (~8-9) is exactly what
    # clipped the last digit — the fix must clear that with room to spare.
    assert ws.column_dimensions["A"].width >= 12
    # "Annual Leave" is 12 chars.
    assert ws.column_dimensions["B"].width >= 14


def test_a_very_long_cell_does_not_blow_out_the_column_past_the_cap():
    original = _xlsx_bytes([["x" * 500]])
    widened = _autofit_xlsx_columns(original)
    wb = load_workbook(io.BytesIO(widened))
    assert wb.active.column_dimensions["A"].width == 60


def test_non_xlsx_bytes_are_returned_unchanged():
    """Legacy binary .xls (and any corrupt file) can't be opened by openpyxl
    — callers already have their own fallback for that, so this must not
    raise, and must hand back the original bytes untouched."""
    garbage = b"not a real spreadsheet"
    assert _autofit_xlsx_columns(garbage) == garbage
